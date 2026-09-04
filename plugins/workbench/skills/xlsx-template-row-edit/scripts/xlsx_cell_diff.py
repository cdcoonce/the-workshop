#!/usr/bin/env python3
# /// script
# requires-python = ">=3.10"
# dependencies = ["openpyxl"]
# ///
"""Diff two .xlsx workbooks cell by cell — value, formula, and style.

    xlsx_cell_diff.py BEFORE.xlsx AFTER.xlsx
    xlsx_cell_diff.py --ref HEAD TEMPLATE.xlsx    # working copy vs a git ref

Exists because a green Python test suite cannot see two whole classes of
damage to a committed report template. ``openpyxl`` never recalculates, so a
formula re-pointed at the wrong rows stays "valid" and every value assertion
reads what the writer wrote rather than what Excel would compute. And styling
is per cell across the full used width, so a style copy that stops at the data
columns silently drops fills and borders in the columns flanking them.

Both failure modes are invisible to a reviewer reading a binary diff, which is
why this compares mechanically instead. Every cell present in either workbook
is compared — never a data-column subset, which would reproduce the second
defect exactly.

Exit contract, mirroring diff(1):
  0  workbooks match
  1  differences found (expected when verifying an intended edit — read them)
  2  could not run (missing file, unreadable workbook, bad git ref)

``openpyxl`` is imported lazily so the comparison logic here stays importable,
and testable, without it.
"""

from __future__ import annotations

import argparse
import subprocess
import sys
import tempfile
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

# Style facets compared per cell, in report order. Each is a plain tuple on
# CellSnapshot so the comparison never touches an openpyxl object.
_STYLE_FIELDS = ("number_format", "font", "fill", "border", "alignment")


@dataclass(frozen=True)
class CellSnapshot:
    """One cell's content and styling, reduced to comparable primitives."""

    value: object
    number_format: str
    font: tuple
    fill: tuple
    border: tuple
    alignment: tuple


@dataclass(frozen=True)
class Difference:
    """One reportable delta at one location."""

    sheet: str
    coordinate: str
    kind: str
    before: object
    after: object


SheetSnapshot = dict[str, CellSnapshot]
WorkbookSnapshot = dict[str, SheetSnapshot]


def _is_formula(value: object) -> bool:
    return isinstance(value, str) and value.startswith("=")


def is_blank(snapshot: CellSnapshot) -> bool:
    """True when a cell carries neither content nor formatting worth diffing.

    Deliberately ignores ``font``: an empty cell's font is inherited noise.
    Fill, border, alignment and number format are NOT ignored — an empty cell
    carrying only a fill is a frame bar, a banner, or a rule, and dropping one
    is the second defect this tool exists to catch.

    Do not reach for ``cell.style`` here. That reports the *named* style, which
    stays ``"Normal"`` for a cell styled directly, so screening on it discards
    every styling-only cell — the exact blind spot being guarded against.
    """
    return (
        snapshot.value is None
        and snapshot.fill == (None, None)
        and snapshot.border == (None, None, None, None)
        and snapshot.alignment == (None, None)
        and snapshot.number_format == "General"
    )


def _sort_key(sheet: str, coordinate: str) -> tuple:
    """Order by sheet, then row, then column — reading order within a sheet."""
    letters = "".join(c for c in coordinate if c.isalpha())
    digits = "".join(c for c in coordinate if c.isdigit())
    column = 0
    for char in letters:
        column = column * 26 + (ord(char.upper()) - ord("A") + 1)
    return (sheet, int(digits or 0), column)


def diff_cells(sheet: str, coordinate: str, before: CellSnapshot, after: CellSnapshot) -> list[Difference]:
    """Compare one cell, reporting every facet that differs.

    A change involving a formula on either side is reported as ``formula``
    rather than ``value``: the distinction is the whole point here, because a
    formula's stored text and its computed result move independently.
    """
    differences: list[Difference] = []
    if before.value != after.value:
        kind = "formula" if _is_formula(before.value) or _is_formula(after.value) else "value"
        differences.append(Difference(sheet, coordinate, kind, before.value, after.value))
    for field in _STYLE_FIELDS:
        old, new = getattr(before, field), getattr(after, field)
        if old != new:
            differences.append(Difference(sheet, coordinate, field, old, new))
    return differences


def diff_workbooks(before: WorkbookSnapshot, after: WorkbookSnapshot) -> list[Difference]:
    """Compare two workbook snapshots, cell by cell, across every sheet.

    A sheet added or removed wholesale is reported once, not once per cell —
    a renamed sheet would otherwise bury real edits under thousands of lines.
    """
    differences: list[Difference] = []
    for sheet in set(before) | set(after):
        if sheet not in before:
            differences.append(Difference(sheet, "", "sheet_added", None, None))
            continue
        if sheet not in after:
            differences.append(Difference(sheet, "", "sheet_removed", None, None))
            continue
        old_cells, new_cells = before[sheet], after[sheet]
        for coordinate in set(old_cells) | set(new_cells):
            if coordinate not in old_cells:
                differences.append(
                    Difference(sheet, coordinate, "added", None, new_cells[coordinate].value)
                )
            elif coordinate not in new_cells:
                differences.append(
                    Difference(sheet, coordinate, "removed", old_cells[coordinate].value, None)
                )
            else:
                differences.extend(
                    diff_cells(sheet, coordinate, old_cells[coordinate], new_cells[coordinate])
                )
    return sorted(differences, key=lambda d: (*_sort_key(d.sheet, d.coordinate), d.kind))


def format_report(differences: Iterable[Difference]) -> str:
    """Render differences as one line each, grouped by location."""
    lines = [
        f"{d.sheet}!{d.coordinate} {d.kind}: {d.before!r} -> {d.after!r}"
        if d.coordinate
        else f"{d.sheet} {d.kind}"
        for d in differences
    ]
    if not lines:
        return "no differences"
    return "\n".join([*lines, "", f"{len(lines)} difference(s)"])


# -- workbook reading (openpyxl imported lazily) ----------------------------


def load_snapshot(path: Path) -> WorkbookSnapshot:
    """Read a workbook into comparable primitives, formulas kept as text."""
    from openpyxl import load_workbook  # noqa: PLC0415 — optional dependency

    workbook = load_workbook(path, data_only=False)
    snapshot: WorkbookSnapshot = {}
    for worksheet in workbook.worksheets:
        cells: SheetSnapshot = {}
        for row in worksheet.iter_rows():
            for cell in row:
                snapshot_cell = CellSnapshot(
                    value=cell.value,
                    number_format=cell.number_format,
                    font=(
                        cell.font.name,
                        cell.font.size,
                        bool(cell.font.bold),
                        bool(cell.font.italic),
                        getattr(cell.font.color, "rgb", None),
                    ),
                    fill=(
                        cell.fill.patternType,
                        getattr(cell.fill.fgColor, "rgb", None),
                    ),
                    border=tuple(
                        getattr(getattr(cell.border, side), "style", None)
                        for side in ("left", "right", "top", "bottom")
                    ),
                    alignment=(cell.alignment.horizontal, cell.alignment.vertical),
                )
                if not is_blank(snapshot_cell):
                    cells[cell.coordinate] = snapshot_cell
        snapshot[worksheet.title] = cells
    return snapshot


def _checkout(ref: str, path: Path, into: Path) -> Path:
    """Materialise ``path`` as of ``ref`` so a working copy can be compared."""
    result = subprocess.run(
        ["git", "show", f"{ref}:{path}"], capture_output=True, check=False
    )
    if result.returncode != 0:
        raise SystemExit(f"cannot read {path} at {ref}: {result.stderr.decode().strip()}")
    target = into / f"{ref.replace('/', '_')}-{path.name}"
    target.write_bytes(result.stdout)
    return target


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    parser.add_argument("before", type=Path, help="baseline workbook, or the edited file with --ref")
    parser.add_argument("after", type=Path, nargs="?", help="workbook to compare against the baseline")
    parser.add_argument("--ref", help="compare the single given workbook against this git ref")
    args = parser.parse_args(argv)

    if bool(args.after) == bool(args.ref):
        parser.error("pass two workbooks, or one workbook with --ref")

    try:
        with tempfile.TemporaryDirectory() as tmp:
            if args.ref:
                before = load_snapshot(_checkout(args.ref, args.before, Path(tmp)))
                after = load_snapshot(args.before)
            else:
                before, after = load_snapshot(args.before), load_snapshot(args.after)
    except SystemExit as exc:
        print(exc, file=sys.stderr)
        return 2
    except (OSError, ValueError, KeyError) as exc:
        print(f"could not read workbook: {exc}", file=sys.stderr)
        return 2

    differences = diff_workbooks(before, after)
    print(format_report(differences))
    return 1 if differences else 0


if __name__ == "__main__":
    raise SystemExit(main())
